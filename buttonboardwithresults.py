# This program reads button press/release events from an Arduino.
# It keeps track of when each button is pressed and released, calculates how long it was pressed,
# and saves all the data in an Excel spreadsheet.

# --- Import necessary modules ---
import serial                  # Lets Python talk to devices connected over USB (like an Arduino)
import time                    # Helps with delays and timing
import threading               # Allows running two things at once (e.g., GUI + reading data)
import tkinter as tk           # Used to build a simple window (GUI)
from tkinter import messagebox # Lets the program show pop-up messages
from openpyxl import Workbook  # Used to create and save Excel files
from datetime import datetime  # Lets us get current time in human-readable format

# --- Serial Setup ---
arduino = serial.Serial(port='COM6', baudrate=9600, timeout=1)  # Connect to Arduino on COM6. Change port if needed.
time.sleep(2)  # Wait 2 seconds for Arduino to reset and be ready to communicate.

# --- Global Data ---
event_log = []  # This list will store all the button events with timestamp and duration
press_times = {}  # Keeps track of when each button was pressed (to later calculate how long it was held)

# --- Serial Reader Thread ---
def read_serial():
    # This function runs in a separate thread and keeps checking for data from the Arduino
    while True:
        if arduino.in_waiting > 0:  # Check if there's data waiting to be read
            line = arduino.readline().decode().strip()  # Read one line, decode it to text, and remove extra spaces
            if not line:
                continue  # If line is empty, skip and keep waiting

            parts = line.split()  # Split the line into two parts (event type and timestamp)
            if len(parts) != 2:
                continue  # Skip if the line doesn't have exactly 2 parts

            event = parts[0]           # This is the event name like "button1_pressed"
            arduino_ms = int(parts[1]) # Time in milliseconds when the event happened on the Arduino
            system_time = datetime.now().strftime('%H:%M:%S.%f')[:-3]  # Current computer time (for logging)

            if "_pressed" in event:  # If this is a press event
                button = event.split("_")[1]         # Extract button number from event string
                press_times[button] = arduino_ms     # Save the time it was pressed
                event_text = f"#{button} - pressed"  # Format the event text for the log
                event_log.append((system_time, event_text, None))  # Save event with no duration yet
                print(f"[{system_time}] {event_text}")  # Show in the terminal

            elif "_released" in event:  # If this is a release event
                button = event.split("_")[1]  # Get the same button number
                if button in press_times:  # Make sure we have a press time saved
                    duration = arduino_ms - press_times[button]  # Calculate how long the button was pressed
                    event_text = f"#{button} - released"         # Format the release event
                    event_log.append((system_time, event_text, duration))  # Save event with duration
                    print(f"[{system_time}] {event_text} (Duration: {duration} ms)")  # Show in terminal
                    del press_times[button]  # Remove the saved press time (cleanup)

# --- Excel Export Function ---
def export_to_excel():
    # This function creates an Excel file from the logged events
    if not event_log:
        messagebox.showwarning("No Data", "No events to export.")  # Show warning if nothing was logged
        return

    wb = Workbook()           # Create a new Excel workbook
    ws = wb.active            # Select the default sheet
    ws.title = "Button Events"  # Name the sheet

    # Set column headers
    ws['A1'] = 'Timestamp'
    ws['B1'] = 'Event'
    ws['C1'] = 'Duration (ms)'

    # Write each event to the sheet
    for idx, (timestamp, event, duration) in enumerate(event_log, start=2):  # Start at row 2 (row 1 has headers)
        ws[f"A{idx}"] = timestamp
        ws[f"B{idx}"] = event
        if duration is not None:
            ws[f"C{idx}"] = duration

    # Create a filename using current date and time
    filename = f"button_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)  # Save the Excel file
    messagebox.showinfo("Export Successful", f"Saved as {filename}")  # Show success message

# --- GUI Setup ---
def start_gui():
    # This sets up a simple window with a message and a button
    root = tk.Tk()  # Create the main window
    root.title("Button Logger")  # Set the window title

    label = tk.Label(root, text="Monitoring button events from Arduino...")  # Create a label
    label.pack(padx=10, pady=10)  # Add spacing around the label

    export_btn = tk.Button(root, text="Print Results", command=export_to_excel)  # Create a button
    export_btn.pack(padx=10, pady=10)  # Add spacing around the button

    root.mainloop()  # Keep the window open and wait for user actions

# --- Start Everything ---

# Start the serial reading in the background so it doesn't block the GUI
serial_thread = threading.Thread(target=read_serial, daemon=True)
serial_thread.start()

# Start the GUI (window)
start_gui()
